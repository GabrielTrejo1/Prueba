from PyQt5 import uic, QtCore
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem
from PyQt5.QtCore import QDate, QDateTime
from openpyxl import load_workbook
import os, sys

class Boleto():
    def __init__(self):
        self.boleto = uic.loadUi("C:/Users/equipo/PycharmProjects/Trabajo-UTN/src/gui/Boleto.ui")
        self.boleto.show()
        self.initGui()

    def initGui(self):
        self.boleto.fechaBoleto.setDate(QDate.currentDate())
        self.boleto.fechaNac1.setDate(QDate.currentDate())
        self.boleto.fechaNac2.setDate(QDate.currentDate())
        self.boleto.fechaauto.setDate(QDate.currentDate())
        self.boleto.fechaauto2.setDate(QDate.currentDate())

        self.boleto.btnGenerar.clicked.connect(self.cargar_datos_excel)  # Cargar datos de plantilla
        self.boleto.btnGenerar_2.clicked.connect(self.guardar_datos)  # Guardar datos de comprador

    def valores(self):
        self.boleto.txtNombre.setText(self.boleto.txtNombre.text())
        self.boleto.txtMarca.setText(self.boleto.txtMarca.text())
        self.boleto.txtModelo.setText(self.boleto.txtModelo.text())

    def cargar_datos_excel(self):
        try:
            plantilla = "C:/Users/equipo/PycharmProjects/Trabajo-UTN/src/templates/plantilla.xlsx"
            salida = "C:/Users/equipo/PycharmProjects/Trabajo-UTN/src/output/boleto.xlsx"  # Cambiar por parametro
            wb = load_workbook(plantilla)
            ws = wb.active
            ##datos comprador
            self.boleto.fechaBoleto.setDate(QDate.fromString(ws['F3'].value, "yyyy-MM-dd"))
            self.boleto.txtNombre.setText(ws['B10'].value or "")
            self.boleto.txtDni1.setText(ws['F10'].value or "")
            self.boleto.fechaNac1.setDate(QDate.fromString(ws['H10'].value, "yyyy-MM-dd" or ""))
            self.boleto.txtConyuge.setText(ws['B11'].value or "")
            self.boleto.txtDni2.setText(ws['F11'].value or "")
            self.boleto.fechaNac2.setDate(QDate.fromString(ws['H11'].value, "yyyy-MM-dd" or ""))
            self.boleto.txtDomicilio.setText(ws['B12'].value or "")
            self.boleto.txtEstCivil.setText(ws['H12'].value or "")
            self.boleto.txtLocalidad.setText(ws['B13'].value or "")
            self.boleto.txtTelefono.setText(ws['B14'].value or "")
            self.boleto.txtActividad.setText(ws['H14'].value or "")
            self.boleto.txtCorreo.setText(ws['C15'].value or "")

            ##datos vehiculos a vender
            self.boleto.txtMarca.setText(ws['B19'].value or "")
            self.boleto.txtModelo.setText(ws['E19'].value or "")
            self.boleto.txtColor.setText(ws['H19'].value or "")
            self.boleto.fechaauto.setDate(QDate.fromString(ws['B20'].value, "yyyy-MM-dd" or ""))
            self.boleto.txtMotor.setText(ws['E20'].value or "")
            self.boleto.txtDominio.setText(ws['B21'].value or "")
            self.boleto.txtChasis.setText(ws['E21'].value or "")
            self.boleto.txtKilometros.setText(ws['H21'].value or "")

            ##Datos vehiculo/s como parte de pago
            self.boleto.txtMarca_dos.setText(ws['B25'].value or "")
            self.boleto.txtmodelo_dos.setText(ws['E25'].value or "")
            self.boleto.txtKm.setText(ws['H25'].value or "")
            self.boleto.fechaauto2.setDate(QDate.fromString(ws['B26'].value, "yyyy-MM-dd" or ""))
            self.boleto.txtMotor_dos.setText(ws['E26'].value or "")
            self.boleto.txtDominio_dos.setText(ws['B27'].value or "")
            self.boleto.txtChasis_dos.setText(ws['E27'].value or "")

            ##datos pago
            self.boleto.txtMonto.setText(ws['D29'].value or "")
            self.boleto.txtContado.setText(ws['B30'].value or "")
            self.boleto.txtCredito.setText(ws['B31'].value or "")
            self.boleto.txtSenia.setText(ws['B32'].value or "")

            ##observaciones
            self.boleto.txtBbservaciones.setText('A41'.value or "")

            wb.save(salida)
            os.startfile(salida)
        except Exception as e:
            (QMessageBox.critical(self.boleto, "Error", f"No se pudo cargar el archivo: {e}"))

    def guardar_datos(self):
        try:
            plantilla = "C:/Users/equipo/PycharmProjects/Trabajo-UTN/src/templates/plantilla.xlsx"
            salida = "C:/Users/equipo/PycharmProjects/Trabajo-UTN/src/output/boleto.xlsx"  # Cambiar por parametro
            wb = load_workbook(plantilla)
            ws = wb.active
            ##datos comprador
            ws['F3'].value = self.boleto.fechaBoleto.date().toString("yyyy-MM-dd")
            ws['B10'].value = self.boleto.txtNombre.text()
            ws['F10'].value = self.boleto.txtDni1.text()
            ws['H10'].value = self.boleto.fechaNac1.date().toString("yyyy-MM-dd")
            ws['B11'].value = self.boleto.txtConyuge.text()
            ws['F11'].value = self.boleto.txtDni2.text()
            ws['H11'].value = self.boleto.fechaNac2.date().toString("yyyy-MM-dd")
            ws['B12'].value = self.boleto.txtDomicilio.text()
            ws['H12'].value = self.boleto.txtEstCivil.text()
            ws['B13'].value = self.boleto.txtLocalidad.text()
            ws['B14'].value = self.boleto.txtTelefono.text()
            ws['H14'].value = self.boleto.txtActividad.text()
            ws['C15'].value = self.boleto.txtCorreo.text()

            ##datos venta vehiculos
            ws['B19'].value = self.boleto.txtMarca.text()
            ws['E19'].value = self.boleto.txtModelo.text()
            ws['H19'].value = self.boleto.txtColor.text()
            ws['B20'].value = self.boleto.fechaauto.date().toString("yyyy")
            ws['E20'].value = self.boleto.txtMotor.text()
            ws['B21'].value = self.boleto.txtDominio.text()
            ws['E21'].value = self.boleto.txtChasis.text()
            ws['H21'].value = self.boleto.txtKilometros.text()

            ##Datos vehiculo/s como parte de pago
            ws['B25'].value = self.boleto.txtMarca_dos.text()
            ws['E25'].value = self.boleto.txtmodelo_dos.text()
            ws['H25'].value = self.boleto.txtKm.text()
            ws['B26'].value = self.boleto.fechaauto2.date().toString("yyyy")
            ws['E26'].value = self.boleto.txtMotor_dos.text()
            ws['B27'].value = self.boleto.txtDominio_dos.text()
            ws['E27'].value = self.boleto.txtChasis_dos.text()

            ##datos pago
            ws['D29'].value = self.boleto.txtMonto.text()
            ws['B30'].value = self.boleto.txtContado.text()
            ws['B31'].value = self.boleto.txtCredito.text()
            ws['B32'].value = self.boleto.txtSenia.text()

            ##observaciones
            ws['A41'].value = self.boleto.txtBbservaciones.text()

            wb.save(salida)
            os.startfile(salida)
        except Exception as e:
            (QMessageBox.critical(self.boleto, "Error", f"No se pudo guardar el archivo: {e}"))
